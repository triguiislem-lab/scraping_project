
#!/usr/bin/env python3
"""
DPM Tunisia exact class/subclass mapper + RCP verifier/downloader.

What this tool does
-------------------
1) Loads the medicine registry workbook exported from DPM.
2) Uses the live DPM class/subclass search pages to scrape the exact therapeutic
   class and subclass for each medicine.
3) Uses the live detail page (fiche) when available to extract the direct RCP link.
4) Verifies candidate RCP links and optionally downloads the PDFs.
5) Writes CSV + Excel outputs with a full manifest.

Expected inputs
---------------
- An Excel workbook with at least a sheet named "Medicines_Enriched"
  or a flat file containing the columns:
  NOM, DOSAGE, FORME, PRESENTATION, NOM GENERIQUE, LABO, PAYS, AMM, DATE AMM, G/P

- A JSON file containing the DPM class/subclass map. This package ships with one.

Install
-------
pip install requests beautifulsoup4 pandas openpyxl pypdf

Run
---
python dpm_tn_live_mapper_and_rcp_tool.py \
  --input-xlsx dpm_tn_enriched_package.xlsx \
  --class-map dpm_tn_class_map.json \
  --output-dir dpm_live_out \
  --mode all \
  --download-rcp

Notes
-----
- This script is designed to run from a normal internet-connected machine.
- It is intentionally conservative with retries and small delays to avoid hammering DPM.
- Because DPM uses legacy forms, the script tries multiple base paths automatically.
"""
from __future__ import annotations

import argparse
import random
import urllib3
import csv
import hashlib
import json
import os
import re
import time
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from pypdf import PdfReader
except Exception:
    PdfReader = None

DEFAULT_BASE_PATHS = [
    "https://dpm.tn/dpm_pharm/medicament",
    "http://dpm.tn/dpm_pharm/medicament",
    "https://www.dpm.tn/dpm_pharm/medicament",
    "http://www.dpm.tn/dpm_pharm/medicament",
    "https://dpm.tn/medicament/humain",
    "https://dpm.tn/medicaments-a-usage-humain",
]
KNOWN_LANDING_PAGES = [
    "https://dpm.tn/medicament/humain/liste-des-medicaments",
    "https://dpm.tn/medicaments-a-usage-humain/liste-des-medicaments",
    "https://dpm.tn/index.php/medicaments-a-usage-humain/liste-des-medicaments",
    "https://dpm.tn/medicament/humain/amms-par-classe-therapeutique",
    "https://dpm.tn/medicaments-a-usage-humain/amms-par-classe-therapeutique",
]
PROBE_FILES = [
    "listclasse.php",
    "listmedic_classe.php",
    "listmedicparnomspec.php",
    "listmedicspec.php",
    "listdci.php",
]
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
)

REGISTRY_COLUMNS = [
    "NOM", "DOSAGE", "FORME", "PRESENTATION", "NOM GENERIQUE",
    "LABO", "PAYS", "AMM", "DATE AMM", "G/P"
]

EXACT_COLUMNS = [
    "EXACT_CLASS_CODE", "EXACT_CLASS_NAME", "EXACT_SUBCLASS_CODE", "EXACT_SUBCLASS_NAME",
    "DPM_LISTING_URL", "DPM_DETAIL_URL",
    "MAP_MATCH_METHOD", "MAP_MATCH_SCORE", "MAP_MATCH_STATUS", "MAP_NOTE",
    "DIRECT_RCP_URL", "RCP_VERIFY_STATUS", "RCP_HTTP_STATUS", "RCP_SOURCE",
    "VERIFIED_RCP_URL", "DOWNLOADED_RCP_FILE", "RCP_SHA256", "RCP_BYTES",
    "RCP_TEXT_PATH", "RCP_TEXT_CHARS", "RCP_EXTRACT_STATUS", "RCP_SECTION_TITLES",
    "RCP_INDICATIONS_TITLE", "RCP_POSOLOGIE_TITLE", "RCP_CONTRE_INDICATIONS_TITLE",
    "LAST_CHECK_UTC"
]

LABEL_ALIASES = {
    "nom": "NOM",
    "nom medicament": "NOM",
    "nom du medicament": "NOM",
    "dénomination": "NOM",
    "denomination": "NOM",
    "specialite": "NOM",
    "spécialité": "NOM",
    "dosage": "DOSAGE",
    "forme": "FORME",
    "présentation": "PRESENTATION",
    "presentation": "PRESENTATION",
    "nom generique": "NOM GENERIQUE",
    "nom générique": "NOM GENERIQUE",
    "dci": "NOM GENERIQUE",
    "labo": "LABO",
    "laboratoire": "LABO",
    "pays": "PAYS",
    "amm": "AMM",
    "date amm": "DATE AMM",
    "classe therapeutique": "EXACT_CLASS_NAME",
    "classe thérapeutique": "EXACT_CLASS_NAME",
    "sous classe": "EXACT_SUBCLASS_NAME",
    "sous-classe": "EXACT_SUBCLASS_NAME",
    "sous classe therapeutique": "EXACT_SUBCLASS_NAME",
    "sous-classe thérapeutique": "EXACT_SUBCLASS_NAME",
    "sous classe thérapeutique": "EXACT_SUBCLASS_NAME",
    "sous-classe therapeutique": "EXACT_SUBCLASS_NAME",
    "indication": "MAP_NOTE",
}

@dataclass
class ListingRecord:
    listing_url: str
    detail_url: Optional[str]
    subclass_code: str
    subclass_name: str
    class_code: str
    class_name: str
    nom: Optional[str] = None
    dosage: Optional[str] = None
    forme: Optional[str] = None
    presentation: Optional[str] = None

@dataclass
class DetailRecord:
    detail_url: str
    fields: Dict[str, str]
    direct_rcp_url: Optional[str]

def normalize_text(value: object) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    s = s.replace("\xa0", " ")
    s = re.sub(r"\s+", " ", s)
    return s

def strip_accents(value: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFKD", value)
        if not unicodedata.combining(c)
    )

def key_text(value: object) -> str:
    s = normalize_text(value)
    s = strip_accents(s).upper()
    s = s.replace("Œ", "OE").replace("Æ", "AE")
    s = re.sub(r"['’`´]", "", s)
    s = re.sub(r"[^A-Z0-9]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def slugify(value: object) -> str:
    s = normalize_text(value)
    s = strip_accents(s).lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s

def dosage_slug(value: object) -> str:
    s = normalize_text(value)
    s = strip_accents(s).lower()
    s = s.replace("µ", "u")
    s = re.sub(r"[^a-z0-9]+", "", s)
    return s

def now_utc() -> str:
    return pd.Timestamp.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")

def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()

def ensure_session(insecure: bool = False) -> requests.Session:
    s = requests.Session()
    s.headers.update({
        "User-Agent": USER_AGENT,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "fr,en;q=0.9",
        "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
        "Referer": "https://dpm.tn/",
        "Origin": "https://dpm.tn",
    })
    s.verify = not insecure
    if insecure:
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    return s

def load_class_map(path: Path) -> Tuple[List[dict], List[dict]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    return data["classes"], data["subclasses"]

def load_registry(path: Path) -> pd.DataFrame:
    if path.suffix.lower() == ".xlsx":
        wb = load_workbook(path, read_only=True, data_only=True)
        preferred = "Medicines_Enriched" if "Medicines_Enriched" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[preferred]
        rows = list(ws.values)
        header = [normalize_text(x) for x in rows[0]]
        df = pd.DataFrame(rows[1:], columns=header)
        wb.close()
    elif path.suffix.lower() in {".csv", ".tsv", ".txt"}:
        sep = "\t" if path.suffix.lower() in {".tsv", ".txt"} else ","
        df = pd.read_csv(path, sep=sep, dtype=str)
    else:
        raise ValueError(f"Unsupported input file: {path}")
    for c in REGISTRY_COLUMNS:
        if c not in df.columns:
            raise ValueError(f"Missing required column: {c}")
    df = df.copy()
    df.insert(0, "ROW_ID", range(1, len(df) + 1))
    for c in REGISTRY_COLUMNS:
        df[c] = df[c].fillna("").map(normalize_text)
    if "RCP_URL_GUESS_GENERIC" not in df.columns:
        df["RCP_URL_GUESS_GENERIC"] = ""
    if "RCP_URL_GUESS_NOM" not in df.columns:
        df["RCP_URL_GUESS_NOM"] = ""
    for c in EXACT_COLUMNS:
        if c not in df.columns:
            df[c] = ""
    df["NOM_KEY"] = df["NOM"].map(key_text)
    df["DOSAGE_KEY"] = df["DOSAGE"].map(key_text)
    df["FORME_KEY"] = df["FORME"].map(key_text)
    df["PRESENTATION_KEY"] = df["PRESENTATION"].map(key_text)
    df["GENERIC_KEY"] = df["NOM GENERIQUE"].map(key_text)
    df["LABO_KEY"] = df["LABO"].map(key_text)
    df["AMM_KEY"] = df["AMM"].map(key_text)
    return df

def build_indices(df: pd.DataFrame) -> Dict[str, dict]:
    return {
        "by_amm": defaultdict(list, {k: v.index.tolist() for k, v in df.groupby("AMM_KEY").groups.items()}),
        "by_nom": defaultdict(list, {k: v.index.tolist() for k, v in df.groupby("NOM_KEY").groups.items()}),
        "by_nom_dose": defaultdict(list, {
            f"{df.loc[i, 'NOM_KEY']}|{df.loc[i, 'DOSAGE_KEY']}": idxs
            for i, idxs in ((idxs[0], idxs) for _, idxs in df.groupby(["NOM_KEY", "DOSAGE_KEY"]).groups.items())
        }),
    }

def try_request(
    session: requests.Session,
    method: str,
    url: str,
    verbose: bool = False,
    **kwargs
) -> Optional[requests.Response]:
    timeout = kwargs.pop("timeout", 30)
    for attempt in range(3):
        try:
            resp = session.request(method, url, timeout=timeout, **kwargs)
            if verbose:
                print(f"[NET] {method.upper()} {url} -> {resp.status_code} ({resp.url})")
            return resp
        except requests.RequestException as exc:
            if verbose:
                print(f"[NET] {method.upper()} {url} -> EXC {type(exc).__name__}: {exc}")
            if attempt == 2:
                return None
            time.sleep(1.5 * (attempt + 1) + random.uniform(0.15, 0.8))
    return None

def _looks_like_legacy_dpm_page(text: str) -> bool:
    t = (text or "").lower()
    markers = [
        "classes therapeutiques",
        "classes thérapeutiques",
        "cod_souscl",
        "cod_classe",
        "listmedic_classe.php",
        "listclasse.php",
        "recherche des médicaments enregistrées par classes thérapeutiques",
    ]
    return any(m in t for m in markers)

def _extract_endpoint_base_from_html(url: str, html: str) -> Optional[str]:
    if not html:
        return None
    for m in re.finditer(r'action=["\']([^"\']*(?:listclasse|listmedic_classe|listmedicparnomspec|listmedicspec)\.php)[^"\']*["\']', html, flags=re.I):
        action = m.group(1)
        full = requests.compat.urljoin(url, action)
        return full.rsplit("/", 1)[0]
    return None

def discover_working_base(session: requests.Session, base_paths: List[str], verbose: bool = False) -> Optional[str]:
    candidate_urls: List[str] = []
    seen = set()

    def add(url: str) -> None:
        if url and url not in seen:
            seen.add(url)
            candidate_urls.append(url)

    for base in base_paths:
        base = base.rstrip("/")
        for probe in PROBE_FILES:
            add(f"{base}/{probe}")

    for landing in KNOWN_LANDING_PAGES:
        add(landing)

    for url in candidate_urls:
        resp = try_request(session, "GET", url, verbose=verbose, allow_redirects=True)
        if resp is None:
            continue

        if resp.status_code == 200:
            html = resp.text or ""
            derived = _extract_endpoint_base_from_html(resp.url, html)
            if derived:
                if verbose:
                    print(f"[DISCOVER] derived endpoint base from form action: {derived}")
                return derived

            if any(resp.url.rstrip("/").lower().endswith("/" + probe.lower()) for probe in PROBE_FILES):
                if verbose:
                    print(f"[DISCOVER] accepted direct endpoint: {resp.url}")
                return resp.url.rsplit("/", 1)[0]

            if _looks_like_legacy_dpm_page(html):
                if verbose:
                    print(f"[DISCOVER] accepted landing page by markers: {resp.url}")
                return resp.url.rsplit("/", 1)[0]

    return None

def parse_listing_records(html: str, listing_url: str, class_code: str, class_name: str, subclass_code: str, subclass_name: str) -> List[ListingRecord]:
    soup = BeautifulSoup(html, "html.parser")
    records: List[ListingRecord] = []
    seen = set()

    # Pattern 1: rows in tables
    for tr in soup.find_all("tr"):
        cells = tr.find_all(["td", "th"])
        if not cells:
            continue
        texts = [normalize_text(c.get_text(" ", strip=True)) for c in cells]
        texts = [t for t in texts if t]
        links = tr.find_all("a", href=True)
        detail_url = None
        for a in links:
            href = a["href"]
            if "fiche" in href.lower() or href.lower().endswith(".php"):
                detail_url = requests.compat.urljoin(listing_url, href)
                break
        if len(texts) >= 1 and (detail_url or len(texts) >= 3):
            nom = texts[0] if texts else None
            dosage = texts[1] if len(texts) > 1 else None
            forme = texts[2] if len(texts) > 2 else None
            presentation = texts[3] if len(texts) > 3 else None
            key = (nom, dosage, forme, presentation, detail_url)
            if key not in seen:
                seen.add(key)
                records.append(ListingRecord(
                    listing_url=listing_url,
                    detail_url=detail_url,
                    subclass_code=subclass_code,
                    subclass_name=subclass_name,
                    class_code=class_code,
                    class_name=class_name,
                    nom=nom, dosage=dosage, forme=forme, presentation=presentation,
                ))

    # Pattern 2: anchor-only pages
    for a in soup.find_all("a", href=True):
        href = a["href"]
        txt = normalize_text(a.get_text(" ", strip=True))
        if "fiche" in href.lower() and txt:
            detail_url = requests.compat.urljoin(listing_url, href)
            key = (txt, None, None, None, detail_url)
            if key not in seen:
                seen.add(key)
                records.append(ListingRecord(
                    listing_url=listing_url,
                    detail_url=detail_url,
                    subclass_code=subclass_code,
                    subclass_name=subclass_name,
                    class_code=class_code,
                    class_name=class_name,
                    nom=txt
                ))
    return records

def parse_detail_record(detail_url: str, html: str) -> DetailRecord:
    soup = BeautifulSoup(html, "html.parser")
    fields: Dict[str, str] = {}

    # table-like label-value parsing
    for tr in soup.find_all("tr"):
        cells = tr.find_all(["td", "th"])
        if len(cells) >= 2:
            label = normalize_text(cells[0].get_text(" ", strip=True)).rstrip(":")
            value = normalize_text(cells[-1].get_text(" ", strip=True))
            label_key = strip_accents(label).lower()
            if label_key in LABEL_ALIASES and value:
                fields[LABEL_ALIASES[label_key]] = value

    # fallback: scan text blobs
    full_text = normalize_text(soup.get_text("\n", strip=True))
    patterns = {
        "NOM": [
            r"specialit[eé][:\s]+(.+?)\n",
            r"nom(?: du)? medicament[:\s]+(.+?)\n",
            r"d[ée]nomination[:\s]+(.+?)\n",
        ],
        "DOSAGE": [r"dosage[:\s]+(.+?)\n"],
        "FORME": [r"forme[:\s]+(.+?)\n"],
        "PRESENTATION": [r"pr[ée]sentation[:\s]+(.+?)\n"],
        "NOM GENERIQUE": [
            r"dci[:\s]+(.+?)\n",
            r"nom g[ée]n[ée]rique[:\s]+(.+?)\n",
        ],
        "LABO": [r"laboratoire[:\s]+(.+?)\n"],
        "PAYS": [r"pays[:\s]+(.+?)\n"],
        "AMM": [r"amm[:\s]+([A-Z0-9]+)\n"],
        "DATE AMM": [r"date amm[:\s]+(.+?)\n"],
        "EXACT_CLASS_NAME": [r"classe th[ée]rapeutique[:\s]+(.+?)\n"],
        "EXACT_SUBCLASS_NAME": [
            r"sous[- ]classe[:\s]+(.+?)\n",
            r"sous[- ]classe th[ée]rapeutique[:\s]+(.+?)\n",
        ],
    }
    for field, plist in patterns.items():
        if fields.get(field):
            continue
        for p in plist:
            m = re.search(p, full_text, flags=re.I)
            if m:
                fields[field] = normalize_text(m.group(1))
                break

    direct_rcp_url = None
    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        href_l = href.lower()
        txt = normalize_text(a.get_text(" ", strip=True)).lower()
        if href_l.endswith(".pdf") or "/images/rcp/" in href_l or "rcp" in txt:
            direct_rcp_url = requests.compat.urljoin(detail_url, href)
            break

    return DetailRecord(detail_url=detail_url, fields=fields, direct_rcp_url=direct_rcp_url)

def score_match(candidate: pd.Series, source_fields: Dict[str, str]) -> int:
    score = 0
    if source_fields.get("AMM") and key_text(candidate["AMM"]) == key_text(source_fields["AMM"]):
        score += 100
    if source_fields.get("NOM") and key_text(candidate["NOM"]) == key_text(source_fields["NOM"]):
        score += 40
    if source_fields.get("DOSAGE") and key_text(candidate["DOSAGE"]) == key_text(source_fields["DOSAGE"]):
        score += 20
    if source_fields.get("FORME") and key_text(candidate["FORME"]) == key_text(source_fields["FORME"]):
        score += 15
    if source_fields.get("PRESENTATION") and key_text(candidate["PRESENTATION"]) == key_text(source_fields["PRESENTATION"]):
        score += 10
    if source_fields.get("NOM GENERIQUE") and key_text(candidate["NOM GENERIQUE"]) == key_text(source_fields["NOM GENERIQUE"]):
        score += 10
    if source_fields.get("LABO") and key_text(candidate["LABO"]) == key_text(source_fields["LABO"]):
        score += 5
    return score

def match_to_registry(df: pd.DataFrame, source_fields: Dict[str, str]) -> Tuple[Optional[int], int, str]:
    candidates = set(df.index.tolist())

    amm = key_text(source_fields.get("AMM", ""))
    if amm:
        subset = df.index[df["AMM_KEY"] == amm].tolist()
        if subset:
            if len(subset) == 1:
                return subset[0], 100, "AMM exact"
            candidates = set(subset)

    nom = key_text(source_fields.get("NOM", ""))
    if nom:
        subset = df.index[df["NOM_KEY"] == nom].tolist()
        if subset:
            if len(candidates) != len(df):
                candidates &= set(subset)
            else:
                candidates = set(subset)

    if not candidates:
        return None, 0, "No candidate"

    scored = []
    for idx in candidates:
        sc = score_match(df.loc[idx], source_fields)
        if sc > 0:
            scored.append((idx, sc))
    if not scored:
        return None, 0, "No scored match"
    scored.sort(key=lambda x: x[1], reverse=True)
    best_idx, best_score = scored[0]
    tied = [x for x in scored if x[1] == best_score]
    if len(tied) == 1:
        return best_idx, best_score, "Scored unique"
    # Prefer presentation tie-break
    if source_fields.get("PRESENTATION"):
        pres = key_text(source_fields["PRESENTATION"])
        exact = [idx for idx, sc in tied if df.loc[idx, "PRESENTATION_KEY"] == pres]
        if len(exact) == 1:
            return exact[0], best_score + 1, "Tie-broken by presentation"
    return None, best_score, f"Ambiguous tie ({len(tied)})"

def fetch_subclass_page(
    session: requests.Session,
    base: str,
    subclass_code: str,
    verbose: bool = False
) -> Tuple[Optional[str], Optional[str]]:
    listing_url = f"{base}/listmedic_classe.php"
    payloads = [
        {"cod_souscl": subclass_code, "submit1": "Continuer"},
        {"cod_souscl": subclass_code},
        {"cod_souscl": subclass_code, "submit": "Continuer"},
    ]
    for payload in payloads:
        resp = try_request(session, "POST", listing_url, data=payload, verbose=verbose)
        if resp is not None and resp.status_code == 200 and resp.text:
            return resp.url, resp.text
    resp = try_request(session, "GET", listing_url, params={"cod_souscl": subclass_code}, verbose=verbose)
    if resp is not None and resp.status_code == 200 and resp.text:
        return resp.url, resp.text
    return None, None

def verify_pdf_url(session: requests.Session, url: str, out_dir: Path, filename_stem: str, download: bool, verbose: bool = False) -> Dict[str, str]:
    result = {
        "RCP_VERIFY_STATUS": "missing",
        "RCP_HTTP_STATUS": "",
        "VERIFIED_RCP_URL": "",
        "DOWNLOADED_RCP_FILE": "",
        "RCP_SHA256": "",
        "RCP_BYTES": "",
    }
    # HEAD can fail on old servers; GET is the reliable fallback
    head = try_request(session, "HEAD", url, allow_redirects=True)
    if head is not None:
        result["RCP_HTTP_STATUS"] = str(head.status_code)
    ok = False
    final_url = url
    if head is not None and head.status_code == 200:
        ctype = (head.headers.get("Content-Type") or "").lower()
        ok = ("pdf" in ctype) or (url.lower().endswith(".pdf"))
        final_url = head.url
    if not ok:
        get = try_request(session, "GET", url, allow_redirects=True, stream=True)
        if get is None:
            return result
        result["RCP_HTTP_STATUS"] = str(get.status_code)
        final_url = get.url
        if get.status_code != 200:
            return result
        first = next(get.iter_content(64), b"")
        content_type = (get.headers.get("Content-Type") or "").lower()
        ok = first.startswith(b"%PDF") or ("pdf" in content_type) or final_url.lower().endswith(".pdf")
        if not ok:
            return result
        if download:
            out_dir.mkdir(parents=True, exist_ok=True)
            path = out_dir / f"{filename_stem}.pdf"
            with open(path, "wb") as f:
                f.write(first)
                for chunk in get.iter_content(1024 * 1024):
                    if chunk:
                        f.write(chunk)
            result["DOWNLOADED_RCP_FILE"] = str(path)
            result["RCP_BYTES"] = str(path.stat().st_size)
            result["RCP_SHA256"] = sha256_of_path(path)
    else:
        if download:
            get = try_request(session, "GET", final_url, allow_redirects=True, stream=True)
            if get is not None and get.status_code == 200:
                out_dir.mkdir(parents=True, exist_ok=True)
                path = out_dir / f"{filename_stem}.pdf"
                with open(path, "wb") as f:
                    for chunk in get.iter_content(1024 * 1024):
                        if chunk:
                            f.write(chunk)
                result["DOWNLOADED_RCP_FILE"] = str(path)
                result["RCP_BYTES"] = str(path.stat().st_size)
                result["RCP_SHA256"] = sha256_of_path(path)
    result["RCP_VERIFY_STATUS"] = "verified"
    result["VERIFIED_RCP_URL"] = final_url
    return result

def sha256_of_path(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()

def build_rcp_candidates(row: pd.Series) -> List[Tuple[str, str]]:
    candidates: List[Tuple[str, str]] = []
    def add(source: str, url: str) -> None:
        url = normalize_text(url)
        if not url:
            return
        if not url.lower().startswith("http"):
            return
        if url not in [u for _, u in candidates]:
            candidates.append((source, url))
    add("detail_page", row.get("DIRECT_RCP_URL", ""))
    add("generic_guess", row.get("RCP_URL_GUESS_GENERIC", ""))
    add("nom_guess", row.get("RCP_URL_GUESS_NOM", ""))
    # extra fallback patterns
    generic_slug = slugify(row.get("NOM GENERIQUE", ""))
    nom_slug = slugify(row.get("NOM", ""))
    dose_slug = re.sub(r"[^a-z0-9]+", "", strip_accents(normalize_text(row.get("DOSAGE", ""))).lower().replace("µ", "u"))
    if generic_slug and dose_slug:
        add("generic_guess_fallback", f"https://dpm.tn/images/rcp/{generic_slug}_{dose_slug}_rcp.pdf")
    if nom_slug and dose_slug:
        add("nom_guess_fallback", f"https://dpm.tn/images/rcp/{nom_slug}_{dose_slug}_rcp.pdf")
    if generic_slug:
        add("generic_no_dose", f"https://dpm.tn/images/rcp/{generic_slug}_rcp.pdf")
    if nom_slug:
        add("nom_no_dose", f"https://dpm.tn/images/rcp/{nom_slug}_rcp.pdf")
    return candidates

def excel_safe_text(value: object, limit: int = 32000) -> str:
    s = normalize_text(value)
    return s if len(s) <= limit else s[: limit - 3] + "..."


def extract_pdf_text(pdf_path: Path) -> Tuple[str, str]:
    if PdfReader is None:
        return "", "pypdf_not_installed"
    try:
        reader = PdfReader(str(pdf_path))
        pages = []
        for page in reader.pages:
            try:
                pages.append(page.extract_text() or "")
            except Exception:
                pages.append("")
        full_text = "\n\n".join(pages)
        full_text = full_text.replace("\\x00", " ")
        full_text = re.sub(r"[ \t]+", " ", full_text)
        full_text = re.sub(r"\n{3,}", "\n\n", full_text)
        return full_text.strip(), "ok"
    except Exception as e:
        return "", f"extract_error:{type(e).__name__}"


_SECTION_VARIANTS = {
    "indications": [
        r"^indications?( therapeutiques?)?$",
        r"^4\.1 indications? therapeutiques?$",
    ],
    "posologie": [
        r"^posologie( et mode d administration)?$",
        r"^mode d administration$",
        r"^4\.2 posologie( et mode d administration)?$",
    ],
    "contre_indications": [
        r"^contre indications?$",
        r"^4\.3 contre indications?$",
    ],
}

_OTHER_HEADING_VARIANTS = [
    r"^composition.*$",
    r"^forme pharmaceutique.*$",
    r"^mises? en garde.*$",
    r"^precautions? d emploi.*$",
    r"^interactions?.*$",
    r"^grossesse.*$",
    r"^allaitement.*$",
    r"^effets? indesirables?.*$",
    r"^surdosage.*$",
    r"^proprietes? pharmacologiques?.*$",
    r"^donnees? pharmaceutiques?.*$",
    r"^titulaire.*$",
    r"^date de .*revis.*$",
    r"^4\.[4-9].*$",
    r"^5\..*$",
    r"^6\..*$",
    r"^7\..*$",
    r"^8\..*$",
    r"^9\..*$",
]


def _normalize_heading_line(line: str) -> str:
    s = strip_accents(normalize_text(line)).lower()
    s = s.replace("'", " ")
    s = re.sub(r"[^a-z0-9. ]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _line_matches_any(line: str, patterns: List[str]) -> bool:
    return any(re.match(p, line, flags=re.I) for p in patterns)


def find_rcp_sections(full_text: str) -> Dict[str, str]:
    lines = [normalize_text(x) for x in full_text.splitlines()]
    norm_lines = [_normalize_heading_line(x) for x in lines]
    result = {
        "RCP_SECTION_TITLES": "",
        "RCP_INDICATIONS_TITLE": "",
        "RCP_INDICATIONS_TEXT": "",
        "RCP_POSOLOGIE_TITLE": "",
        "RCP_POSOLOGIE_TEXT": "",
        "RCP_CONTRE_INDICATIONS_TITLE": "",
        "RCP_CONTRE_INDICATIONS_TEXT": "",
    }
    found_titles = []

    all_heading_patterns = []
    for pats in _SECTION_VARIANTS.values():
        all_heading_patterns.extend(pats)
    all_heading_patterns.extend(_OTHER_HEADING_VARIANTS)

    for key, pats in _SECTION_VARIANTS.items():
        start_idx = None
        for i, nline in enumerate(norm_lines):
            if _line_matches_any(nline, pats):
                start_idx = i
                break
        if start_idx is None:
            continue
        title = lines[start_idx]
        found_titles.append(title)
        end_idx = len(lines)
        for j in range(start_idx + 1, len(lines)):
            nline = norm_lines[j]
            if not nline:
                continue
            if _line_matches_any(nline, all_heading_patterns):
                end_idx = j
                break
        body = "\n".join(x for x in lines[start_idx + 1:end_idx] if normalize_text(x)).strip()
        if key == "indications":
            result["RCP_INDICATIONS_TITLE"] = title
            result["RCP_INDICATIONS_TEXT"] = body
        elif key == "posologie":
            result["RCP_POSOLOGIE_TITLE"] = title
            result["RCP_POSOLOGIE_TEXT"] = body
        elif key == "contre_indications":
            result["RCP_CONTRE_INDICATIONS_TITLE"] = title
            result["RCP_CONTRE_INDICATIONS_TEXT"] = body
    result["RCP_SECTION_TITLES"] = " | ".join(found_titles)
    return result


def extract_and_save_rcp_text(pdf_path: Path, text_dir: Path, file_stem: str) -> Dict[str, str]:
    result = {
        "RCP_TEXT_PATH": "",
        "RCP_TEXT_CHARS": "",
        "RCP_EXTRACT_STATUS": "not_extracted",
        "RCP_SECTION_TITLES": "",
        "RCP_INDICATIONS_TITLE": "",
        "RCP_INDICATIONS_TEXT": "",
        "RCP_POSOLOGIE_TITLE": "",
        "RCP_POSOLOGIE_TEXT": "",
        "RCP_CONTRE_INDICATIONS_TITLE": "",
        "RCP_CONTRE_INDICATIONS_TEXT": "",
        "RCP_FULL_TEXT": "",
        "RCP_FULL_TEXT_PREVIEW": "",
    }
    if not pdf_path.exists():
        result["RCP_EXTRACT_STATUS"] = "pdf_missing"
        return result
    full_text, status = extract_pdf_text(pdf_path)
    result["RCP_EXTRACT_STATUS"] = status
    if status != "ok":
        return result
    text_dir.mkdir(parents=True, exist_ok=True)
    text_path = text_dir / f"{file_stem}.txt"
    text_path.write_text(full_text, encoding="utf-8")
    result["RCP_TEXT_PATH"] = str(text_path)
    result["RCP_TEXT_CHARS"] = str(len(full_text))
    result["RCP_FULL_TEXT"] = full_text
    result["RCP_FULL_TEXT_PREVIEW"] = excel_safe_text(full_text, limit=12000)
    sections = find_rcp_sections(full_text)
    result.update(sections)
    return result


def write_checkpoint_csvs(output_dir: Path, registry: pd.DataFrame, subclass_logs: List[dict], rcp_rows: Optional[List[dict]] = None, rcp_text_rows: Optional[List[dict]] = None) -> None:
    checkpoint_registry = output_dir / "checkpoint_medicines_exact_mapped.csv"
    checkpoint_subclasses = output_dir / "checkpoint_subclass_scrape_log.csv"
    registry.drop(columns=[c for c in registry.columns if c.endswith("_KEY")]).to_csv(checkpoint_registry, index=False, encoding="utf-8-sig")
    pd.DataFrame(subclass_logs).to_csv(checkpoint_subclasses, index=False, encoding="utf-8-sig")
    if rcp_rows is not None:
        pd.DataFrame(rcp_rows).to_csv(output_dir / "checkpoint_rcp_manifest.csv", index=False, encoding="utf-8-sig")
    if rcp_text_rows is not None:
        pd.DataFrame(rcp_text_rows).to_csv(output_dir / "checkpoint_rcp_text_extracts.csv", index=False, encoding="utf-8-sig")


def autosize_worksheet(ws) -> None:
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            v = str(cell.value)
            widths[cell.column] = min(max(widths.get(cell.column, 0), len(v) + 2), 50)
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def write_output_excel(df: pd.DataFrame, subclasses_df: pd.DataFrame, rcp_df: pd.DataFrame, rcp_text_df: pd.DataFrame, output_path: Path) -> None:
    rcp_text_excel = rcp_text_df.copy() if not rcp_text_df.empty else pd.DataFrame(columns=["ROW_ID"])
    if not rcp_text_excel.empty:
        for col in [
            "RCP_FULL_TEXT", "RCP_INDICATIONS_TEXT", "RCP_POSOLOGIE_TEXT", "RCP_CONTRE_INDICATIONS_TEXT",
            "RCP_FULL_TEXT_PREVIEW",
        ]:
            if col in rcp_text_excel.columns:
                rcp_text_excel[col] = rcp_text_excel[col].map(lambda x: excel_safe_text(x, 32000))
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.drop(columns=[c for c in df.columns if c.endswith("_KEY")]).to_excel(writer, sheet_name="Medicines_Mapped", index=False)
        subclasses_df.to_excel(writer, sheet_name="Subclass_Scrape_Log", index=False)
        rcp_df.to_excel(writer, sheet_name="RCP_Manifest", index=False)
        rcp_text_excel.to_excel(writer, sheet_name="RCP_Text_Extracts", index=False)
        summary = pd.DataFrame([
            ["mapped_rows", int((df["MAP_MATCH_STATUS"] == "mapped").sum())],
            ["unmapped_rows", int((df["MAP_MATCH_STATUS"] != "mapped").sum())],
            ["verified_rcp", int((df["RCP_VERIFY_STATUS"] == "verified").sum())],
            ["downloaded_rcp", int((df["DOWNLOADED_RCP_FILE"] != "").sum())],
            ["extracted_rcp_text", int((df["RCP_EXTRACT_STATUS"] == "ok").sum())],
        ], columns=["metric", "value"])
        summary.to_excel(writer, sheet_name="Summary", index=False)

    wb = load_workbook(output_path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        autosize_worksheet(ws)
    wb.save(output_path)


def run_mapper(args: argparse.Namespace) -> None:
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    session = ensure_session(insecure=args.insecure)
    base = discover_working_base(session, args.base_paths, verbose=args.verbose_net)
    if not base:
        raise RuntimeError("Could not reach any DPM base path. Run this on a normal internet-connected machine.")
    print(f"[INFO] Using base path: {base}")

    classes, subclasses = load_class_map(Path(args.class_map))
    registry = load_registry(Path(args.input_xlsx))

    subclass_logs = []
    detail_cache: Dict[str, DetailRecord] = {}

    for subclass in subclasses:
        code = subclass["SUBCLASS_CODE"]
        print(f"[MAP] Fetching subclass {code} - {subclass['SUBCLASS_NAME']}")
        listing_url, html = fetch_subclass_page(session, base, code, verbose=args.verbose_net)
        if not html:
            subclass_logs.append({
                **subclass,
                "listing_url": listing_url or "",
                "status": "fetch_failed",
                "record_count": 0,
            })
            time.sleep(args.delay)
            continue

        listing_records = parse_listing_records(
            html=html,
            listing_url=listing_url or f"{base}/listmedic_classe.php",
            class_code=subclass["CLASS_CODE"],
            class_name=subclass["CLASS_NAME"],
            subclass_code=code,
            subclass_name=subclass["SUBCLASS_NAME"],
        )
        subclass_logs.append({
            **subclass,
            "listing_url": listing_url or "",
            "status": "ok",
            "record_count": len(listing_records),
        })

        for rec in listing_records:
            source_fields = {
                "NOM": rec.nom or "",
                "DOSAGE": rec.dosage or "",
                "FORME": rec.forme or "",
                "PRESENTATION": rec.presentation or "",
                "EXACT_CLASS_NAME": rec.class_name,
                "EXACT_SUBCLASS_NAME": rec.subclass_name,
            }
            direct_rcp_url = ""
            if rec.detail_url:
                if rec.detail_url not in detail_cache:
                    resp = try_request(session, "GET", rec.detail_url, verbose=args.verbose_net)
                    if resp is not None and resp.status_code == 200 and resp.text:
                        detail_cache[rec.detail_url] = parse_detail_record(rec.detail_url, resp.text)
                    else:
                        detail_cache[rec.detail_url] = DetailRecord(rec.detail_url, {}, None)
                    time.sleep(args.delay)
                detail = detail_cache[rec.detail_url]
                source_fields.update({k: v for k, v in detail.fields.items() if v})
                direct_rcp_url = detail.direct_rcp_url or ""

            matched_idx, score, method = match_to_registry(registry, source_fields)
            if matched_idx is None:
                continue

            # Only overwrite if this match is better than what is already stored
            prev_score = pd.to_numeric(pd.Series([registry.at[matched_idx, "MAP_MATCH_SCORE"]]), errors="coerce").fillna(-1).iloc[0]
            if score < prev_score:
                continue

            registry.at[matched_idx, "EXACT_CLASS_CODE"] = rec.class_code
            registry.at[matched_idx, "EXACT_CLASS_NAME"] = source_fields.get("EXACT_CLASS_NAME", rec.class_name)
            registry.at[matched_idx, "EXACT_SUBCLASS_CODE"] = rec.subclass_code
            registry.at[matched_idx, "EXACT_SUBCLASS_NAME"] = source_fields.get("EXACT_SUBCLASS_NAME", rec.subclass_name)
            registry.at[matched_idx, "DPM_LISTING_URL"] = rec.listing_url
            registry.at[matched_idx, "DPM_DETAIL_URL"] = rec.detail_url or ""
            registry.at[matched_idx, "DIRECT_RCP_URL"] = direct_rcp_url
            registry.at[matched_idx, "MAP_MATCH_METHOD"] = method
            registry.at[matched_idx, "MAP_MATCH_SCORE"] = score
            registry.at[matched_idx, "MAP_MATCH_STATUS"] = "mapped"
            registry.at[matched_idx, "LAST_CHECK_UTC"] = now_utc()

        time.sleep(args.delay)
        if args.checkpoint_every_subclass:
            write_checkpoint_csvs(output_dir, registry, subclass_logs)
            print(f"[CHECKPOINT] Saved partial mapping after subclass {code}")

    # Mark unmapped
    registry.loc[registry["MAP_MATCH_STATUS"] == "", "MAP_MATCH_STATUS"] = "unmapped"
    registry.loc[registry["LAST_CHECK_UTC"] == "", "LAST_CHECK_UTC"] = now_utc()

    write_checkpoint_csvs(output_dir, registry, subclass_logs)
    subclasses_df = pd.DataFrame(subclass_logs)
    mapped_csv = output_dir / "medicines_exact_mapped.csv"
    subclass_csv = output_dir / "subclass_scrape_log.csv"
    registry.drop(columns=[c for c in registry.columns if c.endswith("_KEY")]).to_csv(mapped_csv, index=False, encoding="utf-8-sig")
    subclasses_df.to_csv(subclass_csv, index=False, encoding="utf-8-sig")

    # Optional RCP phase
    rcp_rows = []
    rcp_text_rows = []
    if args.mode in {"all", "rcp"}:
        pdf_dir = output_dir / "rcp_pdfs"
        text_dir = output_dir / "rcp_texts"
        for idx, row in registry.iterrows():
            candidates = build_rcp_candidates(row)
            if not candidates:
                continue
            file_stem = re.sub(r"[^a-z0-9_]+", "_", slugify(row["AMM"] or row["NOM"]))[:120]
            verified = None
            chosen_source = ""
            for source, url in candidates:
                result = verify_pdf_url(session, url, pdf_dir, file_stem, download=args.download_rcp, verbose=args.verbose_net)
                if result["RCP_VERIFY_STATUS"] == "verified":
                    verified = result
                    chosen_source = source
                    break
                time.sleep(max(args.delay / 2, 0.1))
            if verified:
                registry.at[idx, "RCP_VERIFY_STATUS"] = verified["RCP_VERIFY_STATUS"]
                registry.at[idx, "RCP_HTTP_STATUS"] = verified["RCP_HTTP_STATUS"]
                registry.at[idx, "VERIFIED_RCP_URL"] = verified["VERIFIED_RCP_URL"]
                registry.at[idx, "DOWNLOADED_RCP_FILE"] = verified["DOWNLOADED_RCP_FILE"]
                registry.at[idx, "RCP_SHA256"] = verified["RCP_SHA256"]
                registry.at[idx, "RCP_BYTES"] = verified["RCP_BYTES"]
                registry.at[idx, "RCP_SOURCE"] = chosen_source
                registry.at[idx, "LAST_CHECK_UTC"] = now_utc()
                manifest_row = {
                    "ROW_ID": row["ROW_ID"],
                    "NOM": row["NOM"],
                    "AMM": row["AMM"],
                    "RCP_SOURCE": chosen_source,
                    **verified
                }
                rcp_rows.append(manifest_row)

                extracted = None
                pdf_path = Path(verified["DOWNLOADED_RCP_FILE"]) if verified.get("DOWNLOADED_RCP_FILE") else None
                if pdf_path and pdf_path.exists():
                    extracted = extract_and_save_rcp_text(pdf_path, text_dir, file_stem)
                    registry.at[idx, "RCP_TEXT_PATH"] = extracted["RCP_TEXT_PATH"]
                    registry.at[idx, "RCP_TEXT_CHARS"] = extracted["RCP_TEXT_CHARS"]
                    registry.at[idx, "RCP_EXTRACT_STATUS"] = extracted["RCP_EXTRACT_STATUS"]
                    registry.at[idx, "RCP_SECTION_TITLES"] = extracted["RCP_SECTION_TITLES"]
                    registry.at[idx, "RCP_INDICATIONS_TITLE"] = extracted["RCP_INDICATIONS_TITLE"]
                    registry.at[idx, "RCP_POSOLOGIE_TITLE"] = extracted["RCP_POSOLOGIE_TITLE"]
                    registry.at[idx, "RCP_CONTRE_INDICATIONS_TITLE"] = extracted["RCP_CONTRE_INDICATIONS_TITLE"]
                    registry.at[idx, "LAST_CHECK_UTC"] = now_utc()
                    rcp_text_rows.append({
                        "ROW_ID": row["ROW_ID"],
                        "NOM": row["NOM"],
                        "AMM": row["AMM"],
                        "VERIFIED_RCP_URL": verified["VERIFIED_RCP_URL"],
                        "DOWNLOADED_RCP_FILE": verified["DOWNLOADED_RCP_FILE"],
                        "RCP_TEXT_PATH": extracted["RCP_TEXT_PATH"],
                        "RCP_TEXT_CHARS": extracted["RCP_TEXT_CHARS"],
                        "RCP_EXTRACT_STATUS": extracted["RCP_EXTRACT_STATUS"],
                        "RCP_SECTION_TITLES": extracted["RCP_SECTION_TITLES"],
                        "RCP_INDICATIONS_TITLE": extracted["RCP_INDICATIONS_TITLE"],
                        "RCP_INDICATIONS_TEXT": extracted["RCP_INDICATIONS_TEXT"],
                        "RCP_POSOLOGIE_TITLE": extracted["RCP_POSOLOGIE_TITLE"],
                        "RCP_POSOLOGIE_TEXT": extracted["RCP_POSOLOGIE_TEXT"],
                        "RCP_CONTRE_INDICATIONS_TITLE": extracted["RCP_CONTRE_INDICATIONS_TITLE"],
                        "RCP_CONTRE_INDICATIONS_TEXT": extracted["RCP_CONTRE_INDICATIONS_TEXT"],
                        "RCP_FULL_TEXT": extracted["RCP_FULL_TEXT"],
                        "RCP_FULL_TEXT_PREVIEW": extracted["RCP_FULL_TEXT_PREVIEW"],
                    })
                elif args.download_rcp:
                    registry.at[idx, "RCP_EXTRACT_STATUS"] = "download_missing"
                else:
                    registry.at[idx, "RCP_EXTRACT_STATUS"] = "download_required_for_text"
            else:
                registry.at[idx, "RCP_VERIFY_STATUS"] = "missing"
                registry.at[idx, "LAST_CHECK_UTC"] = now_utc()
                rcp_rows.append({
                    "ROW_ID": row["ROW_ID"],
                    "NOM": row["NOM"],
                    "AMM": row["AMM"],
                    "RCP_SOURCE": "",
                    "RCP_VERIFY_STATUS": "missing",
                    "RCP_HTTP_STATUS": "",
                    "VERIFIED_RCP_URL": "",
                    "DOWNLOADED_RCP_FILE": "",
                    "RCP_SHA256": "",
                    "RCP_BYTES": "",
                })
            if args.checkpoint_every_rcp and len(rcp_rows) % args.checkpoint_every_rcp == 0:
                write_checkpoint_csvs(output_dir, registry, subclass_logs, rcp_rows, rcp_text_rows)
                print(f"[CHECKPOINT] Saved partial RCP data after {len(rcp_rows)} rows")
            time.sleep(max(args.delay / 2, 0.1))

    write_checkpoint_csvs(output_dir, registry, subclass_logs, rcp_rows, rcp_text_rows)
    rcp_df = pd.DataFrame(rcp_rows)
    rcp_text_df = pd.DataFrame(rcp_text_rows)
    if not rcp_df.empty:
        rcp_df.to_csv(output_dir / "rcp_manifest.csv", index=False, encoding="utf-8-sig")
    if not rcp_text_df.empty:
        rcp_text_df.to_csv(output_dir / "rcp_text_extracts.csv", index=False, encoding="utf-8-sig")

    excel_out = output_dir / "dpm_tn_live_verified.xlsx"
    write_output_excel(
        registry,
        subclasses_df,
        rcp_df if not rcp_df.empty else pd.DataFrame(columns=["ROW_ID"]),
        rcp_text_df if not rcp_text_df.empty else pd.DataFrame(columns=["ROW_ID"]),
        excel_out,
    )
    print(f"[DONE] CSV and Excel outputs written to: {output_dir}")

def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="DPM Tunisia exact class/subclass mapper + RCP verifier/downloader")
    p.add_argument("--input-xlsx", required=True, help="Input registry workbook")
    p.add_argument("--class-map", required=True, help="JSON file containing class/subclass map")
    p.add_argument("--output-dir", required=True, help="Directory for outputs")
    p.add_argument("--mode", choices=["mapping", "rcp", "all"], default="all",
                   help="mapping = class/subclass only, rcp = RCP verification only, all = both")
    p.add_argument("--download-rcp", action="store_true", help="Download verified RCP PDFs")
    p.add_argument("--delay", type=float, default=1.5, help="Delay between requests in seconds")
    p.add_argument("--base-path", dest="base_paths", action="append", default=[],
                   help="Optional DPM base path; can be passed multiple times")
    p.add_argument("--insecure", action="store_true",
                   help="Disable TLS certificate verification for legacy server issues")
    p.add_argument("--verbose-net", action="store_true",
                   help="Print network probing diagnostics")
    p.add_argument("--checkpoint-every-subclass", action="store_true", default=True,
                   help="Write checkpoint CSV outputs after each subclass")
    p.add_argument("--checkpoint-every-rcp", type=int, default=25,
                   help="Write checkpoint CSV outputs every N RCP rows (default: 25)")
    return p

def main() -> None:
    parser = build_parser()
    args = parser.parse_args()
    if not args.base_paths:
        args.base_paths = DEFAULT_BASE_PATHS
    run_mapper(args)

if __name__ == "__main__":
    main()
